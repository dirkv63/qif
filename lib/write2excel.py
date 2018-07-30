"""
This module implements a class to create Excel Workbooks.
"""

import logging
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.exceptions import IllegalCharacterError


class Write2Excel:
    """
    This class consolidates the functions to write the data to excel workbook a nice format.
    """

    def __init__(self):
        """
        Create the workbook object and set the initial sheet.
        """
        self.wb = Workbook()
        self.current_sheet = self.wb.active
        self.rowcnt = 0

    def init_sheet(self, title):
        """
        Initialize the worksheet.

        :param title: mandatory title for the worksheet

        :return: worksheet object
        """
        if self.current_sheet.title == 'Sheet':
            # This is the initial sheet, re-assign title
            self.current_sheet.title = title
        else:
            # There are sheets already, create a new one
            self.current_sheet = self.wb.create_sheet(title=title)
        # Reset rowcnt for new sheet in same workbook
        self.rowcnt = 0
        return self.current_sheet

    def close_workbook(self, filename):
        """
        This method will finalize the current sheet and close the file.
        :param filename:
        :return:
        """
        self.wb.save(filename=filename)
        return

    def write_content(self, content):
        """
        This method will dump the content in an excel worksheet. The content is a list of dictionaries with results.

        :param content:

        :return:
        """
        add_title = True
        for rec in content:
            if add_title:
                self.rowcnt += 1
                colcnt = 0
                for field in rec:
                    colcnt += 1
                    self.current_sheet[rc2a1(self.rowcnt, colcnt)] = field
                add_title = False
            self.rowcnt += 1
            colcnt = 0
            for field in rec:
                colcnt += 1
                try:
                    self.current_sheet[rc2a1(self.rowcnt, colcnt)] = rec[field]
                except IllegalCharacterError:
                    if isinstance(rec[field], bytes):
                        if rec[field] == b'\x00':
                            val = 0
                        elif rec[field] == b'\x01':
                            val = 1
                        else:
                            logging.error("Unexpected byte value: {b}".format(b=rec[field]))
                            val = -1
                        self.current_sheet[rc2a1(self.rowcnt, colcnt)] = val
                    else:
                        logging.error("The value for {f} was not accepted: {s}".format(s=rec[field], f=field))
        return


def rc2a1(row=None, col=None):
    """
    This function converts a (row, column) pair (R1C1 notation) to the A1 string notation for excel. The column number
    is converted to the character(s), the row number is appended to the column string.

    :param row: Row number (1 .. )

    :param col: Column number (1 .. )

    :return: A1 Notation (column-row) (e.g. 'BF19745')
    """
    return "{0}{1}".format(get_column_letter(col), row)
